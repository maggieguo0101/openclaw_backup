from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "\u95e8\u5e97\u57fa\u7840\u4fe1\u606f"

# Header style
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", start_color="2E75B6")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Row fill alternating
fill_white = PatternFill("solid", start_color="FFFFFF")
fill_gray = PatternFill("solid", start_color="F2F2F2")

headers = [
    "\u95e8\u5e97\u540d\u79f0",           # 门店名称
    "\u95e8\u5e97ID",                       # 门店ID
    "\u70b9\u8bc4\u5546\u6237ID",          # 点评商户ID
    "\u5e02/\u533a",                        # 市/区
    "\u884c\u653f\u533a",                   # 行政区
    "\u5546\u5708",                         # 商圈
    "\u95e8\u5e97\u5730\u5740",            # 门店地址
    "\u4e00\u7ea7\u54c1\u7c7b",            # 一级品类
    "\u4e8c\u7ea7\u54c1\u7c7b",            # 二级品类
    "\u7f8e\u56e2\u661f\u7ea7",            # 美团星级
    "\u70b9\u8bc4\u661f\u7ea7",            # 点评星级
    "\u5f53\u524d\u724c\u7ea7",            # 当前牌级
    "\u7d2f\u8ba1\u8bc4\u4ef7\u6570",     # 累计评价数
    "\u8425\u4e1a\u65f6\u95f4",            # 营业时间
    "\u8425\u4e1a\u72b6\u6001",            # 营业状态
    "\u5165\u9a7b\u65f6\u95f4",            # 入驻时间
    "\u5546\u6237\u901a\u7b7e\u7ea6\u65e5\u671f",   # 商户通签约日期
    "\u5546\u6237\u901a\u5230\u671f\u65f6\u95f4",   # 商户通到期时间
    "\u8fd114\u8fd1\u8fd030\u5929\u8bbf\u95ee\u6b21\u6570",  # 近30天访问次数
    "\u8fd114\u8fd1\u8fd030\u5929\u66d9\u5149\u6b21\u6570",  # 近30天曝光次数
    "\u8fd114\u8fd1\u8fd030\u5929\u4e0b\u5355\u5238\u6570",  # 近30天下单券数
    "\u672c\u6708\u8bbf\u95ee\u65b0\u5ba2\u6570",            # 本月访问新客数
    "\u672c\u6708\u8bbf\u95ee\u65b0\u5ba2\u5360\u6bd4",      # 本月访问新客占比
    "\u7ecf\u8425\u8bc4\u5206",                              # 经营评分
    "\u5728\u7ebf\u56e2\u5355\u6570\u91cf",                  # 在线团单数量
    "\u95e8\u5e97\u6807\u7b7e",                              # 门店标签
    "\u7b7e\u7ea6\u9500\u552e",                              # 签约销售
    "\u8fd0\u8425\u9500\u552e",                              # 运营销售
    "\u662f\u5426\u5efa\u7acb\u4f01\u5fae\u95e8\u5e97\u7fa4",  # 是否建立企微门店群
    "\u5ba2\u8d44\u6570",                                    # 客资数
    "\u5907\u6ce8",                                          # 备注
]

ws.row_dimensions[1].height = 40

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

# Data rows: 3 stores
# Conflict resolution:
# Store 1 (亮丽照相馆): city conflict - doc says "东莞" in basic info and "东莞" in city name field -> use 东莞
#   营业状态=0 means 未营业 -> 已停业
# Store 2 (够硬男士理发公司): city conflict - "未知板块"里 city=阳东区, but 门店商圈=湖滨商业街
#   门店所在城市=阳东区 (实际是广东阳江市阳东区), city name confirmed 阳东区
# Store 3 (热炼·健身馆): 容桂细滘居委会文华路65号4-5楼, 顺德区容桂

stores = [
    # Store 1: 亮丽照相馆 (摄影快照, 断约门店, 已停业)
    [
        "\u4eae\u4e3d\u7167\u76f8\u9986",      # 门店名称
        "1015065583856262",                       # 门店ID
        "635881174",                              # 点评商户ID
        "\u4e1c\u839e",                           # 市/区
        "\u5858\u53a6\u9547",                     # 行政区
        "\u5858\u53a6\u9547",                     # 商圈
        "\u65b0\u9633\u8def\u4e2d\u4e0e\u6797\u4e1c\u8def\u4ea4\u53c9\u53e3\u531780\u7c73",  # 门店地址
        "\u751f\u6d3b\u670d\u52a1",               # 一级品类
        "\u5feb\u7167",                           # 二级品类
        3.1,                                      # 美团星级
        "-",                                      # 点评星级
        "\u6682\u65e0\u7b49\u7ea7",               # 当前牌级
        "-",                                      # 累计评价数
        "\u5468\u4e00\u81f3\u5468\u65e5 10:00-21:00",  # 营业时间
        "\u5df2\u505c\u4e1a",                     # 营业状态（0=未营业）
        "2020-05-14",                             # 入驻时间
        "-",                                      # 商户通签约日期
        "-",                                      # 商户通到期时间
        8,                                        # 近30天访问次数
        "-",                                      # 近30天曝光次数
        "-",                                      # 近30天下单券数
        "-",                                      # 本月访问新客数
        "-",                                      # 本月访问新客占比
        "-",                                      # 经营评分
        0,                                        # 在线团单数量
        "\u65ad\u7ea6180\u5929",                  # 门店标签
        "-",                                      # 签约销售
        "-",                                      # 运营销售（王晓慧负责所有门店运营）
        "\u5df2\u5efa\u8054",                     # 是否建立企微门店群
        1590,                                     # 客资数
        "\u65ad\u7ea6\u95e8\u5e97\uff0c\u8fd1\u4e00\u5e74\u9500\u552e\u8ddf\u8fdb4\u6b21\uff1b\u5373\u5c06\u6536\u56de",  # 备注
    ],
    # Store 2: 够硬男士理发公司 (美发, 阳东区)
    [
        "\u591f\u786c\u7537\u58eb\u7406\u53d1\u516c\u53f8",  # 门店名称
        "-",                                      # 门店ID（文档未提供）
        "-",                                      # 点评商户ID（文档未提供）
        "\u9633\u6c5f\u5e02\u9633\u4e1c\u533a",  # 市/区
        "\u9633\u4e1c\u533a",                     # 行政区
        "\u6e56\u6ee8\u5546\u4e1a\u8857",         # 商圈
        "\u4e1c\u5347\u8def22\u53f7",             # 门店地址
        "\u4e3d\u4eba",                           # 一级品类
        "\u7f8e\u53d1",                           # 二级品类
        3.8,                                      # 美团星级
        0.0,                                      # 点评星级
        "\u94dc\u724c",                           # 当前牌级
        243,                                      # 累计评价数
        "\u5468\u4e00\u81f3\u5468\u65e5 13:00-22:00",  # 营业时间
        "\u6b63\u5e38\u8425\u4e1a",               # 营业状态
        "-",                                      # 入驻时间
        "2025-08-15",                             # 商户通签约日期
        "2026-08-14",                             # 商户通到期时间
        3204,                                     # 近30天访问次数（此字段来自第三段数据，与门店描述匹配）
        12916,                                    # 近30天曝光次数
        88,                                       # 近30天下单券数
        47,                                       # 本月访问新客数
        "87.04%",                                 # 本月访问新客占比
        36.0,                                     # 经营评分（当前）
        6,                                        # 在线团单数量
        "-",                                      # 门店标签
        "\u8c22\u5fd7\u9e4f",                     # 签约销售
        "\u738b\u6653\u6167",                     # 运营销售
        "\u5df2\u5efa\u8054",                     # 是否建立企微门店群
        "-",                                      # 客资数
        "\u6709\u5f85\u5b8c\u6210\u56e2\u5355\u4efb\u52a1\uff1a1.\u4fee\u6539\u6d17\u526a\u5439\u56e2\u5355\u4ef7\u683c\u81f3\u2264\u00a529.9\uff1b2.\u65b0\u589e\u5168\u5934\u67d3\u53d1\u56e2\u5355\uff08\u2264\u00a5199\uff09",  # 备注
    ],
    # Store 3: 热炼·健身馆 (运动健身, 顺德容桂)
    [
        "\u70ed\u70bc\u00b7\u5065\u8eab\u9986",   # 门店名称
        "-",                                       # 门店ID
        "-",                                       # 点评商户ID
        "\u4f5b\u5c71\u5e02\u987a\u5fb7\u533a",   # 市/区
        "\u5bb9\u6842\u9547",                      # 行政区
        "\u5bb9\u6842\u7ec6\u6ed9",                # 商圈
        "\u5bb9\u6842\u7ec6\u6ed9\u5c45\u59d4\u4f1a\u6587\u534e\u8def65\u53f74-5\u697c",  # 门店地址
        "\u8fd0\u52a8\u5065\u8eab",                # 一级品类
        "\u5065\u8eab\u4e2d\u5fc3",                # 二级品类
        4.4,                                       # 美团星级
        4.5,                                       # 点评星级
        "\u94dc\u724c",                            # 当前牌级
        243,                                       # 累计评价数
        "\u5468\u4e00\u81f3\u5468\u65e5 09:00-22:00",  # 营业时间
        "\u6b63\u5e38\u8425\u4e1a",                # 营业状态
        "-",                                       # 入驻时间
        "2022-10-10",                              # 商户通签约日期
        "2026-10-10",                              # 商户通到期时间
        3204,                                      # 近30天访问次数
        12916,                                     # 近30天曝光次数
        88,                                        # 近30天下单券数
        47,                                        # 本月访问新客数
        "87.04%",                                  # 本月访问新客占比
        66.2,                                      # 经营评分
        2,                                         # 在线团单数量（包月卡+私教卡）
        "-",                                       # 门店标签
        "\u6ED5\u8D85",                            # 签约销售（滕超）
        "\u738b\u6653\u6167",                      # 运营销售
        "-",                                       # 是否建立企微门店群
        "-",                                       # 客资数
        "\u5168\u6c11\u8fd0\u52a8\u5927\u4fc3\u6d3b\u52a8\u671f\u95f4\uff08\u62db\u5546\uff09\uff1b\u9996\u63a8\u56e2\u5355\uff1a\u5305\u6708\u573a\u5730\u81ea\u7531\u8bad\u7ec3\u5361298.8\u5143\u2192253.9\u5143",  # 备注
    ],
]

data_font = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row_idx, store in enumerate(stores, start=2):
    fill = fill_white if row_idx % 2 == 0 else fill_gray
    ws.row_dimensions[row_idx].height = 30
    for col_idx, val in enumerate(store, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = border
        if col_idx in [1, 7, 8, 9, 14, 31]:
            cell.alignment = left_align
        else:
            cell.alignment = center_align

# Column widths
col_widths = [22,20,14,16,12,14,28,12,12,10,10,12,12,22,12,12,16,16,14,14,14,14,14,10,12,14,12,12,14,10,40]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header
ws.freeze_panes = "A2"

out = "/root/.openclaw/workspace/mock_store_info.xlsx"
wb.save(out)
print(out)
